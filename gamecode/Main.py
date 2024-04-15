import math
import os

import pygame,random,time,sys
from os import path
import openpyxl
from GameSprites import *

#Directories I may use:
img_dir = path.join(path.dirname(__file__), 'img')
aud_dir = path.join(path.dirname(__file__), 'audio')

#constants
WIDTH = 800
HEIGHT = 800
FPS = 60


#Colors I may use:
WHITE = (255, 255, 255)
ALMOST_WHITE = (250, 220, 255)
GRAY1 = (205, 205, 205)
GRAY2 = (155, 155, 155)
GRAY3 = (105, 105, 105)
GRAY3A = (55, 55, 55)
GRAY4 = (40, 40, 40)
GRAY5 = (20, 20, 20)
BLACK = (0, 0, 0)
RED = (255, 0, 0)
GREEN = (0, 255, 0)
BLUE = (0, 0, 255)
YELLOW = (255, 255, 0)
PINK = (255, 20, 147)
CYAN = (50, 50, 255)

enviromental_sprites = pygame.sprite.Group()
global_username = ""
Score = 0
difficulty= ""

class spawn_enemy():
    def __init__(self, type, lane, group, bulletgroup):
        match lane:
            case -3:
                self.spawnx = 100
            case -2:
                self.spawnx = 200
            case -1:
                self.spawnx = 300
            case 0:
                self.spawnx = 400
            case 1:
                self.spawnx = 500
            case 2:
                self.spawnx = 600
            case 3:
                self.spawnx = 700

        match type:
            case "Zaku":
                temp1 = Zaku(self.spawnx, -20)
                group.add(temp1)
            case "Gelgoog":
                temp1 = Gelgoog(self.spawnx, -20, Player1)
                group.add(temp1)
            case "Gouf":
                temp1 = Gouf(self.spawnx, -20, bulletgroup)
                group.add(temp1)
            case "Acguy":
                temp1 = Acguy(self.spawnx, -20, bulletgroup)
                group.add(temp1)


class sqaudren():
    def __init__(self, type, lane, group, bulletgroup):
        match lane:
            case -3:
                self.spawnx = 100
            case -2:
                self.spawnx = 200
            case -1:
                self.spawnx = 300
            case 0:
                self.spawnx = 400
            case 1:
                self.spawnx = 500
            case 2:
                self.spawnx = 600
            case 3:
                self.spawnx = 700

        match type:
            case "Zaku":
                temp1 = Zaku(self.spawnx, 0)
                temp2 = Zaku(self.spawnx - 50, -30)
                temp3 = Zaku(self.spawnx + 50, -30)
                temp4 = Zaku(self.spawnx, -60)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
            case "Zaku-wing":
                temp1 = Zaku(self.spawnx, 0)
                temp2 = Zaku(self.spawnx - 80, -30)
                temp3 = Zaku(self.spawnx + 80, -30)
                temp4 = Zaku(self.spawnx - 160, -30)
                temp5 = Zaku(self.spawnx + 160, -30)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
            case "Zaku-line":
                temp1 = Zaku(self.spawnx - 150, 0)
                temp2 = Zaku(self.spawnx - 50, 0)
                temp3 = Zaku(self.spawnx + 50, 0)
                temp4 = Zaku(self.spawnx + 150, 0)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
            case "Zaku-arrowhead":
                temp1 = Zaku(self.spawnx - 100, 0)
                temp2 = Zaku(self.spawnx, 0)
                temp3 = Zaku(self.spawnx + 100, 0)
                temp4 = Zaku(self.spawnx, -60)
                temp5 = Zaku(self.spawnx, -120)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
            case "GZ":
                temp1 = Gelgoog(self.spawnx + 100, 0, Player1)
                temp2 = Zaku(self.spawnx, 0)
                temp3 = Zaku(self.spawnx + 200, 0)
                temp4 = Zaku(self.spawnx + 50, 0)
                temp5 = Zaku(self.spawnx + 150, 0)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
            case "AZ":
                temp1 = Acguy(self.spawnx + 100, 0, bulletgroup)
                temp2 = Zaku(self.spawnx, 0)
                temp3 = Zaku(self.spawnx + 200, 0)
                temp4 = Zaku(self.spawnx + 50, 0)
                temp5 = Zaku(self.spawnx + 150, 0)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
            case "GZG":
                temp1 = Zaku(self.spawnx - 100, 0)
                temp2 = Zaku(self.spawnx-50, 0)
                temp3 = Gouf(self.spawnx, 0, bulletgroup)
                temp4 = Zaku(self.spawnx + 50, 0)
                temp5 = Zaku(self.spawnx + 100, 0)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
            case "AZG":
                temp1 = Acguy(self.spawnx, -20, bulletgroup)
                temp2 = Zaku(self.spawnx-50, 0)
                temp3 = Gouf(self.spawnx, 0, bulletgroup)
                temp4 = Zaku(self.spawnx + 50, 0)
                temp5 = Zaku(self.spawnx + 100, 0)
                temp6 = Zaku(self.spawnx - 100, 0)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)
                group.add(temp4)
                group.add(temp5)
                group.add(temp6)
            case "Acguy-defensive":
                temp1 = Acguy(self.spawnx, -20, bulletgroup)
                temp2 = Gelgoog(self.spawnx + 50, 0, Player1)
                temp3 = Gelgoog(self.spawnx - 50, 0, Player1)
                group.add(temp1)
                group.add(temp2)
                group.add(temp3)


def draw_text(font, text, x, y):
    text_surface = font.render(text, True, WHITE)
    text_rect = text_surface.get_rect()
    text_rect.center = (x,y)
    screen.blit(text_surface,text_rect)
def draw_cursor (font,x,y,x2,y2):
    draw_text(font, "*", x,y)
    draw_text(font, "*", x2, y2)

class cursor():
    def __init__(self,font):
        self.font = font
        self.x = 250
        self.x2= 550
        self.y = 500
        self.y2 = self.y
        self.text = "*"
        text_surface = font.render(self.text, True, WHITE)
        text_rect = text_surface.get_rect()
        text_rect.center = (self.x, self.y)
        screen.blit(text_surface, text_rect)

        text2_surface = font.render(self.text, True, WHITE)
        text2_rect = text2_surface.get_rect()
        text2_rect.center = (self.x, self.y)
        screen.blit(text2_surface, text2_rect)
    def move (self, x, x2, y):
        self.x = x


def create_account(username, password):
    wb = openpyxl.load_workbook('PlayerLogin.xlsx')
    ws = wb.active
    ws.append([username, password])
    wb.save('PlayerLogin.xlsx')


def check_credentials(username, password):
    wb = openpyxl.load_workbook('PlayerLogin.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

pygame.init()
pygame.font.init()
my_font = pygame.font.Font('img/Vermin Vibes 1989.ttf', 30)
menu_font = pygame.font.Font('img/Vermin Vibes 1989.ttf', 50)
hud_font = pygame.font.Font('img/Vermin Vibes 1989.ttf', 20)
pygame.mixer.init()
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption('MSG:COSMIC_VANGUARD')

logo = pygame.sprite.Sprite()
logo.image = pygame.image.load("img/Logo.png").convert_alpha()
logo.rect = logo.image.get_rect()
logo.rect.center = [400,250]
pxl = pygame.sprite.Sprite()
pxl.image = pygame.image.load("img/Gundampxl.png").convert_alpha()
pxl.rect = pxl.image.get_rect()
pxl.rect.center = (520, 400)

Zpxl = pygame.sprite.Sprite()
Zpxl.image = pygame.image.load("img/zakuArt.png").convert_alpha()
Zpxl.rect = Zpxl.image.get_rect()
Zpxl.rect.center = (150, 350)

RXPXL = pygame.sprite.Sprite()
RXPXL.image = pygame.image.load("img/RXPXL.png").convert_alpha()
RXPXL.rect = RXPXL.image.get_rect()
RXPXL.rect.center = (200, 350)

VGPXL = pygame.sprite.Sprite()
VGPXL.image = pygame.image.load("img/VGPXL.png").convert_alpha()
VGPXL.rect = VGPXL.image.get_rect()
VGPXL.rect.center = (600, 350)



CS = pygame.sprite.Sprite()
CS.image = pygame.image.load("img/controls_screen.png").convert_alpha()
CS.rect = VGPXL.image.get_rect()
CS.rect.center = (160,200)

starsG = pygame.sprite.Group()
starSprite1 = stars(400,400)
starSprite2 = stars(400, -624)
starsG.add(starSprite1)
starsG.add(starSprite2)
clock = pygame.time.Clock()
bullets = pygame.sprite.Group()
PlayerBullets = pygame.sprite.Group()
enemies=pygame.sprite.Group()
enemyGelgoogs = pygame.sprite.Group()
enemyBullets = pygame.sprite.Group()
qabillist = pygame.sprite.Group()
swordslash = pygame.sprite.Group()
healthBarBox=pygame.Rect(0,0,300,20)
Player_img = pygame.image.load(path.join(img_dir,'Gundam-b.png')).convert()
Player_img.set_colorkey(BLACK)
bkg_img = pygame.image.load(path.join(img_dir, 'SpaceP.png')).convert()
dth_img = pygame.image.load(path.join(img_dir, 'dthscr.png')).convert()
win_img = pygame.image.load(path.join(img_dir, 'winscreen.png')).convert()
HUD_img = pygame.image.load(path.join(img_dir, 'NewHUD.png')).convert()
HUD_img.set_colorkey(WHITE)
HUD_img.set_alpha(50)
HUD = Background(HUD_img)
crack = pygame.image.load(path.join(img_dir, 'screen break.png')).convert_alpha()
crack.set_alpha(50)
crackbkg = Background(crack)
background = Background(bkg_img)
Player1 = Player(Player_img,"VG")
explosions = pygame.sprite.Group()
pygame.mixer.init()
boss = pygame.sprite.Group()
shoot_sfx = pygame.mixer.Sound("audio/beam-rifle-shot.mp3")
shoot_sfx.set_volume(.05)
sword_sfx = pygame.mixer.Sound("audio/sword-schwing-40520.mp3")
death_sfx = pygame.mixer.Sound("audio/gundam-deathSFX.mp3")
death_sfx.set_volume(.04)
swordE = energyBar(25, 700)
healthB = healthBar(25, 750)
grenadeB = grenadeBar(675, 750)
grenadeB.energy = 3
grenades = pygame.sprite.Group()
SE = pygame.sprite.Group()
SE.add(swordE)
SE.add(healthB)
SE.add(grenadeB)
dl = dashlight(700,700)
enviromental_sprites.add(dl)

def charSel():
    running = True
    selected = 1
    while running:
        keys = pygame.key.get_pressed()
        screen.blit(bkg_img, (0, 0))
        background.draw(screen)
        screen.blit(VGPXL.image, VGPXL.rect)
        screen.blit(RXPXL.image, RXPXL.rect)
        pygame.draw.rect(screen, BLUE, VGPXL.rect, 3)
        pygame.draw.rect(screen, BLUE, RXPXL.rect, 3)
        if selected ==1:
            pygame.draw.rect(screen, GREEN, RXPXL.rect, 5)
        if selected ==2:
            pygame.draw.rect(screen, GREEN, VGPXL.rect, 5)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if keys[pygame.K_RIGHT]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_LEFT]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_SPACE]:
                if (selected == 1):
                    Player1 = Player(Player_img,"RX")
                    menu(Player1)
                if (selected == 2):
                    Player1 = Player(Player_img,"VG")
                    menu(Player1)

        draw_text(menu_font, 'RX-78-2 Gundam', 200, 600)
        draw_text(menu_font, 'RX-93 v Gundam', 600, 600)

        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()

def controls():
    running = True
    while running:
        keys = pygame.key.get_pressed()
        screen.fill(BLACK)
        screen.blit(bkg_img, (0,0))
        screen.blit(CS.image, CS.rect)
        background.draw(screen)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
        if keys[pygame.K_ESCAPE]:
            menu(Player1)
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()



def menu (P1, username=global_username):
    username = username
    running = True
    selected = 1
    Player1 = P1
    pygame.mixer.music.load("audio/caraaa.mp3")
    pygame.mixer.music.play(100)
    pygame.mixer.music.set_volume(.05)
    while running:

        keys = pygame.key.get_pressed()
        screen.fill(BLACK)
        screen.blit(bkg_img, (0,0))
        background.draw(screen)
        screen.blit(pxl.image, pxl.rect)
        screen.blit(Zpxl.image, Zpxl.rect)
        screen.blit(logo.image, logo.rect)


        if (selected == 1):
            draw_cursor(menu_font, 250, 500, 550, 500)
        if (selected == 2):
            draw_cursor(menu_font, 220, 550, 575, 550)
        if (selected == 3):
            draw_cursor(menu_font, 325, 600, 475, 600)
        if (selected == 4):
            draw_cursor(menu_font, 270, 650, 525, 650)
        if (selected == 5):
            draw_cursor(menu_font, 200, 700, 590, 700)
        if (selected == 6):
            draw_cursor(menu_font, 335, 750, 460, 750)
        draw_text(menu_font, 'Start Game', 400, 500)
        draw_text(menu_font, 'Phase Select', 400, 550)
        draw_text(menu_font, 'Login', 400, 600)
        draw_text(menu_font, 'Controls', 400, 650)
        draw_text(menu_font, 'Gundam Select', 400, 700)
        draw_text(menu_font, 'Quit', 400, 750)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if keys[pygame.K_DOWN]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =3
                    pygame.display.update()
                elif (selected == 3):
                    selected =4
                    pygame.display.update()
                elif (selected == 4):
                    selected =5
                    pygame.display.update()
                elif (selected == 5):
                    selected =6
                    pygame.display.update()
                elif (selected == 6):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_UP]:
                if (selected == 1):
                    selected =6
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
                elif (selected == 3):
                    selected =2
                    pygame.display.update()
                elif (selected == 4):
                    selected =3
                    pygame.display.update()
                elif (selected == 5):
                    selected =4
                    pygame.display.update()
                elif (selected == 6):
                    selected =5
                    pygame.display.update()
            if keys[pygame.K_SPACE]:
                if (selected == 1):
                    diffpicker(P1)
                if (selected == 2):
                    phasepicker(P1)
                if (selected == 3):
                    login()
                if (selected == 4):
                    controls()
                if (selected ==5):
                    charSel()
                if (selected ==6):
                    running = False
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()


def diffpicker (P1, username=global_username, Phase=1):
    username = username
    running = True
    selected = 1
    Player1 = P1
    while running:
        keys = pygame.key.get_pressed()
        screen.fill(BLACK)
        screen.blit(bkg_img, (0,0))
        background.draw(screen)
        screen.blit(pxl.image, pxl.rect)
        screen.blit(Zpxl.image, Zpxl.rect)
        screen.blit(logo.image, logo.rect)


        if (selected == 1):
            draw_cursor(menu_font, 325, 500, 475, 500)
        if (selected == 2):
            draw_cursor(menu_font, 325, 550, 475, 550)
        draw_text(menu_font, 'EASY', 400, 500)
        draw_text(menu_font, 'HARD', 400, 550)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if keys[pygame.K_DOWN]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_UP]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_SPACE]:
                global difficulty
                if selected == 1:
                    difficulty = "EASY"
                    game(P1, Phase)
                elif selected == 2:
                    difficulty = "HARD"
                    game(P1, Phase)
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()


def phasepicker (P1, username=global_username):
    username = username
    running = True
    selected = 1
    Player1 = P1
    while running:
        keys = pygame.key.get_pressed()
        screen.fill(BLACK)
        screen.blit(bkg_img, (0,0))
        background.draw(screen)
        screen.blit(pxl.image, pxl.rect)
        screen.blit(Zpxl.image, Zpxl.rect)
        screen.blit(logo.image, logo.rect)


        if (selected == 1):
            draw_cursor(menu_font, 300, 500, 500, 500)
        if (selected == 2):
            draw_cursor(menu_font, 295, 550, 505, 550)
        if (selected == 3):
            draw_cursor(menu_font, 295, 600, 505, 600)
        if (selected == 4):
            draw_cursor(menu_font, 325, 650, 475, 650)
        draw_text(menu_font, 'Phase 1', 400, 500)
        draw_text(menu_font, 'Phase 2', 400, 550)
        draw_text(menu_font, 'Phase 3', 400, 600)
        draw_text(menu_font, 'Demo', 400, 650)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if keys[pygame.K_DOWN]:
                if (selected == 1):
                    selected =2
                    pygame.display.update()
                elif (selected == 2):
                    selected =3
                    pygame.display.update()
                elif (selected == 3):
                    selected =4
                    pygame.display.update()
                elif (selected == 4):
                    selected =1
                    pygame.display.update()
            if keys[pygame.K_UP]:
                if (selected == 1):
                    selected =4
                    pygame.display.update()
                elif (selected == 2):
                    selected =1
                    pygame.display.update()
                elif (selected == 3):
                    selected = 2
                    pygame.display.update()
                elif (selected == 4):
                    selected =3
                    pygame.display.update()
            if keys[pygame.K_SPACE]:
                global difficulty
                if selected == 1:
                    diffpicker(P1,global_username,1)
                elif selected == 2:
                    diffpicker(P1,global_username,2)
                elif selected == 3:
                    diffpicker(P1,global_username,3)
                elif selected == 4:
                    game (P1, 0)
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()
class Button:
    def __init__(self, x, y, width, height, text, color, hover_color, callback):
        self.rect = pygame.Rect(x, y, width, height)
        self.color = color
        self.hover_color = hover_color
        self.text = text
        self.font = pygame.font.Font(None, 32)
        self.callback = callback
        self.hovered = False

    def draw(self, screen):
        color = self.hover_color if self.hovered else self.color
        pygame.draw.rect(screen, color, self.rect)
        text_surface = self.font.render(self.text, True, (255, 255, 255))
        text_rect = text_surface.get_rect(center=self.rect.center)
        screen.blit(text_surface, text_rect)

    def handle_event(self, event):
        if event.type == pygame.MOUSEMOTION:
            self.hovered = self.rect.collidepoint(event.pos)
        elif event.type == pygame.MOUSEBUTTONDOWN:
            if self.hovered:
                self.callback()

class TextInputBox:
    def __init__(self, x, y, width, height, text=''):
        self.rect = pygame.Rect(x, y, width, height)
        self.color = pygame.Color('lightskyblue3')
        self.text = text
        self.font = pygame.font.Font(None, 32)
        self.txt_surface = self.font.render(text, True, self.color)
        self.active = False

    def handle_event(self, event):
        if event.type == pygame.MOUSEBUTTONDOWN:

            if self.rect.collidepoint(event.pos):
                self.active = not self.active
            else:
                self.active = False
            self.color = pygame.Color('dodgerblue2') if self.active else pygame.Color('lightskyblue3')
        if event.type == pygame.KEYDOWN:
            if self.active:
                if event.key == pygame.K_RETURN:
                    return self.text
                elif event.key == pygame.K_BACKSPACE:
                    self.text = self.text[:-1]
                else:
                    self.text += event.unicode
                self.txt_surface = self.font.render(self.text, True, self.color)
        return None

    def update(self):
        width = max(200, self.txt_surface.get_width() + 10)
        self.rect.w = width

    def draw(self, screen):
        screen.blit(self.txt_surface, (self.rect.x + 5, self.rect.y + 5))
        pygame.draw.rect(screen, self.color, self.rect, 2)
def login():
    def submit_credentials(username):
        global global_username
        if username.strip():
            wb = openpyxl.load_workbook('PlayerLogin.xlsx')
            ws = wb.active
            next_row = 1
            while ws.cell(row=next_row, column=1).value:
                next_row += 1
            ws.cell(row=next_row, column=1).value = username.strip()
            wb.save('PlayerLogin.xlsx')
            global_username = username.strip()
        else:
            draw_text(menu_font, "Must fill out the Username field", 500, 200)

    username_box = TextInputBox(300, 200, 140, 32)


    button_x = 300
    button_y = 250
    button_width = 140
    button_height = 50


    button = Button(button_x, button_y, button_width, button_height, "Submit", (0, 255, 0), (0, 200, 0),
                    lambda: [submit_credentials(username_box.text), menu(Player1, username_box.text)])

    input_boxes = [username_box]

    running = True
    while running:
        screen.fill((0, 0, 0))
        draw_text(menu_font, "Account Management", 400, 100)
        draw_text(menu_font, "Username:", 200, 200)

        for box in input_boxes:
            box.update()
            box.draw(screen)

        button.draw(screen)

        pygame.display.flip()

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    submit_credentials(username_box.text)
                    username = username_box.text
                    menu(Player1, username)
            for box in input_boxes:
                box.handle_event(event)
            button.handle_event(event)




def death_screen():
    running = True
    space_pressed = False
    while running:
        screen.blit(dth_img, (-100, 0))
        draw_text(menu_font, 'press esc to go back to menu', 400, 750)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE and not space_pressed:
                    running = False
                    menu(Player1)

                    space_pressed = True
        pygame.display.flip()
        clock.tick(60)

def win_screen():
    running = True
    space_pressed = False
    while running:
        screen.blit(win_img, (0, 0))
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE and not space_pressed:
                    running = False
                    menu(Player1)

                    space_pressed = True
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
    exit()
def game(P1=Player1, Phase = 1):
    level_selected = False
    Player1 = P1
    PlayerBullets = pygame.sprite.Group()
    score = 0
    done = False
    if difficulty == "HARD":
        multiplier = 2
    else:
        multiplier = 1
    username = global_username
    if Player1.type == "RX":
        swordE.timer = 100
    Char = CharNovaZakuII(50, 50,Player1, enviromental_sprites)
    Ze = Zeong(Player1)
    bossPhase = False
    enviromental_sprites.add(qabillist)
    random_spawn_rate_1 = random.randint(4000, 8000)
    random_spawn_rate_2 = random.randint(4000, 7000)
    random_spawn_rate_3 = random.randint(5000, 6000)
    random_sqaud_rate_1 = random.randint(6000, 9000)
    random_sqaud_rate_2 = random.randint(5000, 8000)
    random_sqaud_rate_3 = random.randint(6000, 8000)
    random_list_1=["Zaku","Gelgoog"]
    random_sqaud_1 = ["Zaku","Zaku-wing","Zaku-line","Zakue-arrowhead","GZ"]
    random_list_2 = ["Zaku","Gelgoog","Acguy"]
    random_sqaud_2 = ["Zaku", "Zaku-wing", "Zaku-line", "Zakue-arrowhead", "GZ", "GZG", "AZ", "Acguy-defesnive"]
    random_list_3 = ["Zaku", "Gelgoog", "Acguy","Gouf"]
    random_sqaud_3 = ["Zaku", "Zaku-wing", "Zaku-line", "Zakue-arrowhead", "GZ", "GZG", "AZ", "Acguy-defesnive", "AZG"]
    last_spawn_time = 0
    last_squad_spawn_time = 0
    DASH_COOLDOWN = 5000
    spawn_rate = None
    spawn_list = None
    sqaud_rate = None
    sqaud_list = None
    last_dash_time = pygame.time.get_ticks()
    Player1.health = 300
    healthB.health = Player1.health
    if os.path.exists('PlayerLogin.xlsx'):
        wb = openpyxl.load_workbook('PlayerLogin.xlsx')
        ws = wb.active
        if ws['A2'].value:
            username = ws['A2'].value
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                if row[0].value == username:
                    current_row = row[0].row
                    break

    else:
        current_row = None

    while not done:
        current_time = pygame.time.get_ticks()
        if current_time >= 60000 and score >= 1000 and Phase == 1:
            Phase =2
        if current_time >= 60000 and score >= 1500 and Phase == 2:
            bossPhase = True
        if current_time >= 60000 and score >= 2000 and Phase == 3:
            bossPhase = True

        if Phase == 1:
            spawn_rate = random_spawn_rate_1
            spawn_list = random_list_1
            sqaud_rate = random_sqaud_rate_1
            sqaud_list = random_sqaud_1
        if Phase == 2:
            spawn_rate = random_spawn_rate_2
            spawn_list = random_list_2
            sqaud_rate = random_sqaud_rate_2
            sqaud_list = random_sqaud_2
        if Phase == 3:
            spawn_rate = random_spawn_rate_3
            spawn_list = random_list_3
            sqaud_rate = random_sqaud_rate_3
            sqaud_list = random_sqaud_3
        if bossPhase == True and Phase == 2:
            boss.add(Char)
        if bossPhase == True and Phase == 3:
            boss.add(Ze)
        if spawn_rate != None:
            if len(enemies) < 15 and current_time-last_spawn_time >= spawn_rate:
                if Phase == 1:
                    if random.randint(1,10) >= 9:
                        spawn_enemy(spawn_list[1], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    else:
                        spawn_enemy(spawn_list[0], random.randint(-3, 3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                if Phase == 2:
                    if 6 <= random.randint(1,10) <= 8 :
                        spawn_enemy(spawn_list[1], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    if random.randint(1,10) >= 9 :
                        spawn_enemy(spawn_list[2], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    else:
                        spawn_enemy(spawn_list[0], random.randint(-3, 3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                if Phase == 3:
                    if 6<= random.randint (1,10) <=7 :
                        spawn_enemy(spawn_list[1], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    if random.randint (1,10) == 8 :
                        spawn_enemy(spawn_list[2], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    if random.randint(1,10) >= 9 :
                        spawn_enemy(spawn_list[3], random.randint(-3,3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
                    else:
                        spawn_enemy(spawn_list[0], random.randint(-3, 3),enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()
        if sqaud_rate != None:
            if len(enemies) < 15 and current_time-last_squad_spawn_time >= sqaud_rate:
                if Phase == 1:
                    if random.randint(1,10) >= 9:
                        sqaudren(sqaud_list[4], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                    else:
                        sqaudren(sqaud_list[random.randint(0,4)], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                if Phase == 2 and bossPhase == False:
                    if 6 < random.randint(1,10) < 9 :
                        spawn_enemy(sqaud_list[random.randint(4,5)], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                    if random.randint(1,10) >= 9 :
                        spawn_enemy(sqaud_list[random.randint(6,7)], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                    else:
                        sqaudren(sqaud_list[random.randint(0, 3)], random.randint(-3, 3), enemies, enemyBullets)
                        last_spawn_time = pygame.time.get_ticks()

                if Phase == 3 and bossPhase == False:
                    if 5 <= random.randint(1,10) <= 7 :
                        spawn_enemy(sqaud_list[random.randint(5,6)], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                    if random.randint(1,10) >= 9 :
                        spawn_enemy(sqaud_list[random.randint(7,8)], random.randint(-3,3),enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()
                    else:
                        sqaudren(sqaud_list[random.randint(0, 3)], random.randint(-3, 3), enemies, enemyBullets)
                        last_squad_spawn_time = pygame.time.get_ticks()

        if username != '':
            if current_row is not None:
                ws.cell(row=current_row, column=2).value = score
                wb.save('PlayerLogin.xlsx')
        keys = pygame.key.get_pressed()
        if current_time - last_dash_time >= DASH_COOLDOWN:
            dash_active = True
            dl.status="on"
        else:
            dash_active = False
            dl.status = "off"
        text_surface = my_font.render(str(score), False, WHITE)
        scoreRect=text_surface.get_rect()
        scoreRect.center=(700,15)
        if keys[pygame.K_w] and Player1.rect.y>0:
            if dash_active and keys[pygame.K_LSHIFT]:
                Player1.rect.y -= 100
                last_dash_time = pygame.time.get_ticks()
            else:
                Player1.rect.y -= 5
        if keys[pygame.K_s] and Player1.rect.y<725:
            if dash_active and keys[pygame.K_LSHIFT]:
                Player1.rect.y += 100
                last_dash_time = pygame.time.get_ticks()
            else:
                Player1.rect.y += 5
        if keys[pygame.K_a] and Player1.rect.x>5:
            if dash_active and keys[pygame.K_LSHIFT]:
                Player1.rect.x -= 100
                last_dash_time = pygame.time.get_ticks()
            else:
                Player1.rect.x -= 5
        if keys[pygame.K_d] and Player1.rect.x<735:
            if dash_active and keys[pygame.K_LSHIFT]:
                Player1.rect.x += 100
                last_dash_time = pygame.time.get_ticks()
            else:
                Player1.rect.x += 5
        if keys[pygame.K_e] and len(swordslash) == 0 and swordE.energy == 4:
            if (Player1.type == "VG"):
                temp1 = Fin (Player1.rect.x - 25, Player1.rect.y, Player1)
                temp2 = Fin(Player1.rect.x + 35, Player1.rect.y,Player1)
                swordslash.add (temp1)
                swordslash.add (temp2)
            else:
                ss = Swordswing(Player1.rect.x-20, Player1.rect.y - 10)
                swordslash.add(ss)
            swordE.energy = 0
        if keys[pygame.K_q] and len(grenades) == 0 and grenadeB.energy >0:
            grenade = Grenade(Player1.rect.x, Player1.rect.y, enemies, enviromental_sprites)
            grenades.add(grenade)
            grenadeB.energy -=1
        if (Phase == 0):
            if keys[pygame.K_1] and len(enemies)<1:
                spawn_enemy("Zaku", random.randint(-3,3),enemies, enemyBullets)
            if keys[pygame.K_2]and len(enemies)<1:
                spawn_enemy("Gelgoog", random.randint(-3,3),enemies, enemyBullets)
            if keys[pygame.K_3]and len(enemies)<1:
                spawn_enemy("Acguy", random.randint(-3,3),enemies, enemyBullets)
            if keys[pygame.K_4]and len(enemies)<1:
                spawn_enemy("Gouf", random.randint(-3,3),enemies, enemyBullets)
            if keys[pygame.K_5]:
                Ze.abilgroup.empty()
                Ze.rect.y = -200
                boss.empty()
                boss.add (Char)
            if keys[pygame.K_6]:
                Char.abilgroup.empty()
                boss.empty()
                boss.add (Ze)


        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True
            elif event.type == pygame.KEYDOWN:
                if event.key == 32:
                    if len(PlayerBullets) < 5:
                        shoot_sfx.play()
                        Player1.shoot(PlayerBullets)
                    if swordslash.__len__()>1 and Player1.type=="VG" and len(PlayerBullets)< 10:
                        temp1.shoot(PlayerBullets)
                        temp2.shoot(PlayerBullets)
        for b in PlayerBullets:
            for e in enemies:
                if pygame.sprite.collide_rect(b,e):
                    if isinstance(e, Gelgoog) and e.stun:
                        e.health -= 100
                        if (e.health <= 0):
                            death_sfx.play()
                            tempx = e.rect.x
                            ex = explosion(tempx, e.rect.y)
                            e.kill()
                            explosions.add(ex)
                            score += e.score
                        b.kill()
                        break
                    elif isinstance(e,Gelgoog) and e.stun == False:
                        b.kill()
                        e.shield-=1
                    else:
                        e.health -=100
                        b.kill()
                        if (e.health <= 0):
                            death_sfx.play()
                            tempx = e.rect.x
                            tempy=e.rect.y
                            ex = explosion(tempx, e.rect.y)
                            if random.randint(0,10) > 9:
                                orb_type = random.choice(['heal', 'energy', 'grenade'])
                                temporb = Orb( tempx+32, tempy, orb_type)
                                enviromental_sprites.add(temporb)
                            explosions.add(ex)
                            score += e.score
                            e.kill()
                        break
        for s in swordslash:
            for b in enemyBullets:
                if pygame.sprite.collide_rect(b,s):
                    b.kill()
            for e in enemies:
                if pygame.sprite.collide_rect(s, e):
                    death_sfx.play()
                    tempx = e.rect.x
                    ex = explosion(tempx, e.rect.y)
                    e.kill()
                    explosions.add(ex)
                    score += e.score
            if boss.__len__()>0:
                for b in boss:
                    if pygame.sprite.collide_rect(s, b):
                        b.health -= 5

        for o in enviromental_sprites:
            if isinstance(o, Orb):
                if pygame.sprite.collide_rect(o, Player1):
                    if o.type == "heal":
                        Player1.health += o.health_value
                        healthB.health += o.health_value
                        if Player1.health > 300:
                            Player1.health = 300
                        if healthB.health > 300:
                            healthB.health = 300
                        o.kill()
                    if o.type == "energy":
                        swordE.energy+=1
                        o.kill()
                    if o.type == "grenade":
                        grenadeB.energy+=1
                        o.kill()
        screen.fill(BLACK)
        screen.blit(bkg_img, background.sprite)
        background.draw(screen)
        explosions.draw(screen)
        explosions.update()
        bullets.draw(screen)
        bullets.update()
        PlayerBullets.draw(screen)
        PlayerBullets.update()
        swordslash.draw(screen)
        swordslash.update()
        enemyBullets.draw(screen)
        enemyBullets.update()
        enviromental_sprites.draw(screen)
        enviromental_sprites.update()
        grenades.update()
        grenades.draw(screen)
        SE.draw(screen)
        SE.update()
        if bossPhase == False:
            draw_text(hud_font, "Phase: "+str(Phase), 50, 10)
        else:
            draw_text(hud_font, "Phase: BOSS", 50, 10)
        if dl.status == "on":
            draw_text(hud_font, "DASH ACTIVE", 720, 740)
        if swordslash.__len__()>1:
            temp1.rect.y = Player1.rect.y
            temp2.rect.y = Player1.rect.y

            if temp1.rect.x != Player1.rect.x - 45:
                temp1.rect.x = Player1.rect.x - 45
            if temp2.rect.x != Player1.rect.x + 50:
                temp2.rect.x = Player1.rect.x + 50
            if pygame.time.get_ticks() - temp1.starttimer > 4999:
                temp1.kill()
                temp2.kill()
            if pygame.time.get_ticks() - temp2.starttimer > 4999:
                temp2.kill()
        starsG.draw(screen)
        starsG.update()
        for b in PlayerBullets:
            b.update()
            if Char.shieldActive and pygame.sprite.collide_rect(Char.shield, b):
                b.kill()
                Char.shield.health -=1
                if Char.shield.health ==0:
                    Char.shield.kill()
            elif not Char.shieldActive and pygame.sprite.collide_rect(Char, b):
                Char.health -= 5
                b.kill()
            elif pygame.sprite.collide_rect(Ze, b):
                Ze.health -= 5
                b.kill()
        for s in swordslash:
            if s.rect.y < 0:
                s.kill()
        for b in enemyBullets:
            b.update()
            if pygame.sprite.collide_rect(b,Player1):
                if difficulty == "HARD" and b.damage != b.damage * 2:
                    b.damage = b.damage * multiplier
                    print(b.damage)
                Player1.health-=b.damage
                b.kill()
        for e in enemies:
            if e.rect.y > 800:
                e.kill()
                score -= e.score
                if score <0:
                    score = 0
            if (random.randint(1,100)<2) and len(enemyBullets)<8:
                spawnx = e.rect.x+15
                if isinstance(e, Gelgoog):
                    b = CresentShot(spawnx, e.rect.y, Player1)
                    enemyBullets.add(b)
                if isinstance(e, Gouf) or isinstance(e, Acguy):
                    pass
                else:
                    b = EnemyBullet(spawnx, e.rect.y+25)
                    enemyBullets.add(b)
        if bossPhase:
            for a in Char.abilgroup:
                if pygame.sprite.collide_rect(a, Player1):
                    Player1.health -= a.damage
                    a.kill()
        enemies.draw(screen)
        healthB.health = Player1.health
        health_percent = Player1.health * 100 // 300
        enemies.update()
        tempRect = Player_img.get_rect()
        tempRect.center = 200, 200
        Player1.draw(screen)
        Player1.update()
        boss.update()
        boss.draw(screen)
        if Char.health <= 0:
            Char.kill()
        if Char.health <= 0 and Phase == 2:
            Char.abilgroup.empty()
            Phase = 3
            bossPhase = False
            Char.kill()
            boss.empty()
        if Ze.health <= 0:
            Ze.abilgroup.empty()
            Ze.kill()
            win_screen()
        Char.abilgroup.draw(screen)
        screen.blit(HUD_img, HUD.sprite)
        HUD.draw(screen)
        Ze.abilgroup.update()
        Ze.abilgroup.draw(screen)
        screen.blit(text_surface, scoreRect)
        if (health_percent < 30):
            screen.blit(crack, crackbkg.sprite)
        if health_percent <= 0:
            Player1.kill()
            enemies.empty()
            bullets.empty()
            boss.empty()
            if username != '' and current_row is not None:
                ws.cell(row=current_row, column=2).value = score
                wb.save('PlayerLogin.xlsx')
            death_screen()
        draw_text(hud_font, 'Ability', 55, 700)
        draw_text(hud_font, 'AP:'+health_percent.__str__()+"%", 75, 750)
        if boss.__len__()>0:
            for b in boss:
                draw_text(hud_font, 'Boss AP:' + b.health.__str__(), 75, 100)
        pygame.display.flip()
        pygame.display.update()
        clock.tick(60)
    pygame.quit()
    exit()
menu(Player1)
